from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
'''from pyscript import Element'''
import os, json, requests, uuid, itertools
from datetime import datetime
from math import floor
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

i,x = 0,1
escolha = []
from static.python.funcAtestado import *

teste()

i = 0
app = Flask(__name__, static_folder='')
app.secret_key = 'chave-secreta'
UPLOAD_FOLDER =  './src/static/uploads/atestados/' if __name__ == '__main__' else './static/uploads/atestados/'
UPLOAD_EQUIPE = './src/static/equipes/' if __name__ == '__main__' else './static/equipes/'
UPLOAD_FOLDER_PDFS = './src/static/uploads/equipes/' if __name__ == '__main__' else './static/uploads/equipes/'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_EQUIPE, exist_ok=True)
os.makedirs(UPLOAD_FOLDER_PDFS, exist_ok=True)
    
@app.route("/")
def index():
    return render_template("index.html"), 200

@app.route('/formAtestado')
def envio():
    return render_template('formAtestado.html')

@app.route('/scrum', methods=['GET', 'POST'])
def scrum():
    equipe_selecionada = None

    if request.method == "POST":
        equipe_selecionada = request.form.get("equipe")
    
    equipes = {}
    try:
        with open(UPLOAD_EQUIPE + 'equipes.txt', 'r', encoding='utf-8') as file:
            for linha in file:
                linha = linha.strip()
                if linha.startswith("Nome da Equipe:"):
                    equipe_atual = linha.replace("Nome da Equipe:", "").strip()
                    equipes[equipe_atual] = {'integrantes': []}
                elif linha.startswith("Nome do Integrante:") and equipe_atual:
                    integrante = linha.replace("Nome do Integrante:", "").strip()
                    equipes[equipe_atual]['integrantes'].append(integrante)
                elif linha.startswith("Avaliacão"):
                    avaliacao_status = linha.split(":", 1)[1].replace('\n', '').strip()
                    equipes[equipe_atual]['statusAvaliacao'] = avaliacao_status.strip().lower() == 'true'
    except FileNotFoundError:
        equipes = {}

    return render_template('scrum.html', equipes=equipes, equipe_selecionada=equipe_selecionada)

@app.route("/integrantesScrum")
def integrantes():
    nomeEquipe = request.args.get("NOME")
    try:
        with open(UPLOAD_EQUIPE+"equipes.txt", "r", encoding='utf-8') as file:
            equipesLinha = file.readlines()

            integrantes=[]
            equipe_achado = False
            for linha in equipesLinha:
                linha = linha.strip()
                if linha.strip().lower() == f"nome da equipe: {nomeEquipe.strip().lower()}":
                    equipe_achado = True
                    continue
                
                if equipe_achado and not linha:
                    equipe_achado = False
                    break
                
                if equipe_achado and linha and not linha.startswith("Avaliacão"):
                    chave_valor = linha.split(":", 1)
                    if len(chave_valor) == 2:
                        chave, valor = chave_valor
                        integrantes.append(valor.strip())
        return jsonify({'integrantes': integrantes}), 200
    except Exception as e:
        return 'Equipe não encontrada', 404
    
@app.route("/avaliacaoStatus")
def avaliacaoStatus():
    nomeEquipe = request.args.get("equipeNome")

    with open(UPLOAD_EQUIPE + "equipes.txt", "r", encoding="utf-8") as file:
        equipe_found = False
        for linha in file:
            if linha.startswith("Nome da Equipe") and nomeEquipe in linha:
                equipe_found = True
                continue

            if not linha.strip():  # corrigido para verificar linha vazia corretamente
                equipe_found = False
                continue

            if linha.startswith("Avaliacão") and equipe_found:
                avalicao_status = linha.split(":", 1)[-1].strip()  # remove \n e espaços
                return jsonify({"avaliacao": avalicao_status == "True"})

    # Se não encontrar, retorna False
    return jsonify({"avaliacao": False})

@app.route('/alterarStatusAvaliacao', methods=['POST'])
def alterarStatusAvaliacao():
    equipeNome = request.form.get("equipe")
    novoStatus = request.form.get("status")
    
    try:
        # Ler todo o arquivo
        with open(UPLOAD_EQUIPE + "equipes.txt", "r", encoding="utf-8") as file:
            linhas = file.readlines()

        equipe_encontrada = False
        for i, linha in enumerate(linhas):
            if linha.strip().startswith("Nome da Equipe:") and equipeNome in linha:
                equipe_encontrada = True
                continue
            
            if equipe_encontrada and linha.strip().startswith("Avaliacão:"):
                linhas[i] = f"Avaliacão: {novoStatus}\n"
                break
            elif not linha.strip():  # Linha vazia, reseta a flag
                equipe_encontrada = False

        # Escrever de volta no arquivo
        with open(UPLOAD_EQUIPE + "equipes.txt", "w", encoding="utf-8") as file:
            file.writelines(linhas)

        flash(f"Status da equipe {equipeNome} alterado para {'Aberta' if novoStatus == 'True' else 'Fechada'}!")
        return redirect(url_for('adminscrum'))

    except Exception as e:
        flash(f"Erro ao alterar status: {str(e)}")
        return render_template("adminAtestado.html")

@app.route('/enviarNotas', methods=['POST'])                   
def enviarNotas():
    dados = request.json
    equipeNome = dados.get("equipeNome")
    avaliador = dados.get("Avaliador")
    avaliado = dados.get("Avaliado")
    notas = dados.get("notas")
    pathArquivo = os.path.join(UPLOAD_EQUIPE, f"{equipeNome}.txt")

    try:
        if os.path.exists(pathArquivo):
            with open(pathArquivo, "r", encoding='utf-8') as fileR:
                arquivo = fileR.readlines()

            substring = f"{avaliador} - {avaliado}"
            existe_avaliacao = any(substring in item for item in arquivo)

            if existe_avaliacao:
                arquivo_reescrito = []

                for linha in arquivo:
                    if linha.startswith(f"{avaliador} - {avaliado}"):
                        nova_linha = f"{avaliador} - {avaliado}: {notas}\n"
                        arquivo_reescrito.append(nova_linha)
                    else:
                        arquivo_reescrito.append(linha)

                with open(pathArquivo, "w", encoding='utf-8') as fileW:
                    fileW.writelines(arquivo_reescrito)

                return jsonify({"mensagem": "Nota Atualizada com sucesso"}), 200

        with open(pathArquivo, "a", encoding='utf-8') as fileA:
            fileA.writelines(f"{avaliador} - {avaliado}: {notas}\n")

        return jsonify({"mensagem": "Nota Enviada com sucesso"}), 200
    except Exception as e:
        return jsonify({"Erro": str(e)}), 500

@app.route("/verificaAvaliacaoCompleta", methods=['GET', 'POST'])
def verificaAvaliacaoCompleta():
    nomeEquipe = request.args.get("NomeEquipe")
    pathArquivo = f"{UPLOAD_EQUIPE}{nomeEquipe}.txt"
    if os.path.exists(pathArquivo):
        with open(pathArquivo, "r", encoding="utf-8") as file:
            arquivo = file.readlines()

        try:
            with app.test_client() as client:
                response = client.get(f'/integrantesScrum?NOME={nomeEquipe}')
                equipeScrum = response.data.decode('utf-8')
        except Exception as e:
            print(e)

        equipeScrum = json.loads(equipeScrum)
        listaIntegrantes_semFormatacao = equipeScrum.get("integrantes", [])
        listaIntegrantes = []

        for integrante in listaIntegrantes_semFormatacao:
            integrante = integrante.split("-")
            integrante = integrante[0]
            listaIntegrantes.append(integrante.strip())

        qntIntegrantes = len(listaIntegrantes)
        avalicao_completo = True
        dictAvaliacao = {}
        for integrante in range(qntIntegrantes):
            qntNotas = 0
            for notasIntegrante in arquivo:
                if notasIntegrante.startswith(f"{listaIntegrantes[integrante]}"):
                    qntNotas += 1

            dictAvaliacao[listaIntegrantes[integrante]] = qntNotas
            
            if qntNotas < qntIntegrantes - 1:
                avalicao_completo = False
                break
        
        return jsonify({"statusAvaliacao": avalicao_completo})
    return jsonify({"Erro": "Avaliacão Fechada"})

@app.route('/adminatestado')
def adminatestado():
    return render_template("adminAtestado.html")

@app.route('/adminscrum', methods=['GET', "POST"])
def ler_equipe():
    equipes = {}
    equipe_atual = None

    if request.method == "POST":
        nome_equipe = request.form.get("nome_equipe")
        num_integrantes = request.form.get("num_integrantes")

        if nome_equipe and num_integrantes and not request.form.get("nome_0"):
            try:
                num_integrantes = int(num_integrantes)
                return render_template("cadastroequipes.html", num_integrantes=num_integrantes)
            except ValueError:
                return redirect(request.url)


    try:
        # ANTES equipes = {'404': ['integrante1, '...']}
        # DEPOIS equipes = {'404': {'integrantes': ['integrante1', 'integrante2'], 'avaliacao': True}, 'teste': []}
        with open(UPLOAD_EQUIPE + 'equipes.txt', 'r', encoding='utf-8') as file:
            for linha in file:
                linha = linha.strip()
                if linha.startswith("Nome da Equipe:"):
                    equipe_atual = linha.replace("Nome da Equipe:", "").strip()
                    equipes[equipe_atual] = {'integrantes': []}
                elif linha.startswith("Nome do Integrante:") and equipe_atual:
                    integrante = linha.replace("Nome do Integrante:", "").strip()
                    equipes[equipe_atual]['integrantes'].append(integrante)

                elif linha.startswith("Avaliacão"):
                    avaliacao_status = linha.split(":", 1)[1].replace('\n', '').strip()
                    equipes[equipe_atual]['statusAvaliacao'] = avaliacao_status.strip().lower() == 'true'
                    
                     
    except FileNotFoundError:
        equipes = {}
    return render_template('adminScrum.html', equipes=equipes)

@app.route('/enviar', methods=['POST'])
def enviar():
    nome = request.form['nome']
    RA = request.form['RA']
    data_i = request.form['data_i']
    data_f = request.form['data_f']
    motivo = request.form['motivo']
    arquivo = request.files['arquivo']
    status = 'Pendente'

    if not arquivo:
        flash('Envie um arquivo válido!')
        return redirect(url_for('index'))
    
    #Calcula dias afastado
    def calcula_validade_atestado(data_i, data_f):
    
        diferenca = abs(datetime.strptime(data_f, "%Y-%m-%d") - datetime.strptime(data_i, "%Y-%m-%d"))
    
        return diferenca.days
    
    # Salvar o arquivo com um nome único para evitar sobrescrita
    file_ext = os.path.splitext(arquivo.filename)[1]  # Extrair a extensão do arquivo
    novo_nome = f"{RA}_{datetime.now().strftime('%d%m%Y%H%M%S')}{file_ext}"  # Novo nome com RA e timestamp
    caminho_arquivo = os.path.join(UPLOAD_FOLDER, novo_nome)
    arquivo.save(caminho_arquivo)

    # Salvar os dados em um arquivo de texto
    with open(UPLOAD_FOLDER+'atestados.txt', 'a', encoding='utf-8') as f:
        f.write(f"Nome: {nome}\nRA do aluno: {RA}\nData Inicial: {data_i}\nData Final: {data_f}\nValidade: {calcula_validade_atestado(data_i, data_f)} dias\nMotivo: {motivo}\nArquivo: {caminho_arquivo}\nStatus: {status}\n\n")

    flash('Atestado enviado com sucesso!')
    return redirect(url_for('ler_txt'))

@app.route('/espera', methods=['GET'])
def ler_txt():
    try:
        with open(UPLOAD_FOLDER+'atestados.txt', 'r', encoding='utf-8') as f:
            linhas = f.readlines()
        #aqui começa a transoformar em um dicionario pra tratar melhor os dados tropa
        atestados = []
        atestado_atual = {}

        for linha in linhas:
            linha = linha.strip()
            
            if not linha:  # verifica linha vazia q mostra separação de registros
                if atestado_atual:  
                    atestados.append(atestado_atual)  # add o atestado na lista
                    atestado_atual = {}  # zera para o próximo
                continue  
            
            chave_valor = linha.split(":", 1)  # Divide na primeira ocorrência de ":"
            if len(chave_valor) == 2:
                chave, valor = chave_valor
                atestado_atual[chave.strip()] = valor.strip()  # Remove espaços extras
        
        if atestado_atual:  # Garante que o último atestado seja salvo
            atestados.append(atestado_atual)

        return render_template('espera.html', atestados=atestados, datetime=datetime)

    except FileNotFoundError:
        return render_template('espera.html', atestados= 'Ih deu ruim')
    
@app.route("/header")
def header():
    return render_template('header.html')

@app.route("/scrum", methods=["GET", "POST"])
def submit():
    media = 0 
    x = len(escolha)
    if request.method == "POST":
        if x < 5:  # Permite no máximo five escolhas
            escolha.append(request.form.get("options"))
        else:
            media = sum(int(i) for i in escolha) / len(escolha)
    return render_template("scrum.html", escolha=escolha, x=x, media=media)

@app.route("/buscar-atestado")
def buscar():
    try:
        with open(UPLOAD_FOLDER + 'atestados.txt', 'r', encoding='utf-8') as file:
            atestados = file.readlines()
            statusLista = []
            RA = request.args.get("RA")
            for i, linha in  enumerate(atestados):
                if linha.startswith("RA do aluno:"):
                    raLinha = linha.split(":")[1].strip()
                
                    if (RA == raLinha):
                        statusLinha = atestados[i+6]
                        if (statusLinha.startswith('Status:')):
                            status = statusLinha.split(":")[1].strip()
                            statusLista.append(status)
            if statusLista:
                print(statusLista)
                return jsonify({"status": statusLista[-1]})
                
            return jsonify({"status": "Não Encontrado"})
    except FileNotFoundError:
        return render_template('Espera.html')        

@app.route("/gestaoat", methods=["GET"])
def gestao():
    try:
        with open(UPLOAD_FOLDER+'atestados.txt', 'r', encoding='utf-8') as f:
            linhas = f.readlines()
        #aqui começa a transoformar em um dicionario pra tratar melhor os dados tropa
        atestados = []
        atestado_atual = {}
        for linha in linhas:
            
            linha = linha.strip()
            
            if not linha:  # verifica linha vazia q mostra separação de registros
                if atestado_atual and atestado_atual['Status'] == 'Pendente': # aqui ele vê se for Pendente, se for ele mostra, senão zera
                    atestados.append(atestado_atual)  # add o atestado na lista
                atestado_atual = {}  # zera para o próximo
                continue  
            
            chave_valor = linha.split(":", 1)  # Divide na primeira ocorrência de ":"
            if len(chave_valor) == 2:
                chave, valor = chave_valor
                atestado_atual[chave.strip()] = valor.strip()  # Remove espaços extras

        if atestado_atual and atestado_atual['Status'] == 'Pendente':# Garante que o último atestado seja salvo
            atestados.append(atestado_atual)
        
        return render_template('GestãoDeAtestados.html', atestados=atestados)

    except FileNotFoundError:
        return render_template('GestãoDeAtestados.html', atestados= 'Ih deu ruim')

@app.route("/alterarStatus")
def alterarStatus():
    url = request.args.get("URL")
    status = request.args.get("status")
    try:
        with open(UPLOAD_FOLDER + 'atestados.txt', 'r', encoding='utf-8') as file:
            atestados = file.readlines()

        indexLinhaStatus = 0
        for i, linha in enumerate(atestados):
            if url in linha:
                indexLinhaStatus = i+1
                break
        
        atestados.pop(indexLinhaStatus)
        atestados.insert(indexLinhaStatus, "Status: "+status+"\n")

        with open(UPLOAD_FOLDER + 'atestados.txt', 'w', encoding='utf-8') as fileW:
                fileW.writelines(atestados)

        return f'Sucesso', 200  

    except Exception as e:
        return f'Erro ao atualizar o status: {str(e)}', 500  
    
@app.route("/cadastroequipes", methods=["POST"])
def cadastro_equipes():
    if request.method == "POST":
        nome_equipe = request.form["nome_equipe"]
        num_integrantes = int(request.form["num_integrantes"])

        if nome_equipe and num_integrantes:
            try:
                with open(UPLOAD_EQUIPE + "equipes.txt", "r", encoding="utf-8") as f:
                    linhas = f.readlines()
                    for l in linhas:
                        if l.strip().lower().startswith("nome da equipe:"):
                            nome_arquivo = l.split(":", 1)[1].strip()
                            if nome_arquivo.lower() == nome_equipe.strip().lower():
                                flash("Equipe existente")
                                return redirect("adminscrum")
            except FileNotFoundError:
                pass


        integrantes = []
        for i in range(num_integrantes):
            nome = request.form.get(f"nome_{i}")
            funcao = request.form.get(f"funcao_{i}")
            if not nome or not funcao:
                flash("Integrante ou função em branco.")
                return redirect("adminscrum")
            integrantes.append((nome, funcao))

        with open(UPLOAD_EQUIPE + "equipes.txt", "a", encoding="utf-8") as f:
            f.write(f"Nome da Equipe: {nome_equipe}\n")
            for nome, funcao in integrantes:
                f.write(f"Nome do Integrante: {nome} - {funcao}\n")
            f.write("Avaliacão: False\n")
            f.write("\n")

        flash("Equipe cadastrada com sucesso!")

    return redirect("adminscrum")

@app.route("/verificaequipe", methods = ["GET","POST"])
def verificarequipe():
    nome_equipe = request.args.get("nome_equipe")
    if nome_equipe:
        try:
            with open(UPLOAD_EQUIPE + "equipes.txt", "r", encoding="utf-8") as f:
                linhas = f.readlines()
                for l in linhas:
                    if l.strip().lower().startswith("nome da equipe:"):
                        nome_arquivo = l.split(":", 1)[1].strip()
                        if nome_arquivo.lower() == nome_equipe.strip().lower():
                            flash("Equipe existente")
                            return "",200
            return "",300
        except FileNotFoundError:
            pass

@app.route("/gerarExcel", methods = ["GET", "POST"])
def genarateExcel():
    nomeEquipe = request.args.get("nomeEquipe")
    pathArquivo = f"{UPLOAD_EQUIPE}{nomeEquipe}.txt"
    excelPath = f"{UPLOAD_FOLDER_PDFS}{nomeEquipe}.xlsx"
    if os.path.exists(pathArquivo):
        with open(pathArquivo, 'r', encoding='utf-8') as file:
            arquivo = file.readlines()

        jsonFile = generateJson(jsonFile=arquivo)

        arquivo = Workbook()
        aba = arquivo.active

        aba.cell(row=1, column=1).value = "Integrante"
        for i, (integrante, info) in enumerate(jsonFile.items(), start=2):
            aba.cell(row=i, column=1).value = integrante
            
            k = 2

            for j, (criterio, media) in enumerate(info['medias'].items(), start=2):
                aba.cell(row=1, column=j).value = criterio
                
                aba.cell(row=i, column=k).value = media
                k+=1

        arquivo.save(excelPath)
        return jsonify({'url': excelPath}), 200
    return jsonify({"error": "Equipe não encontrada"})

@app.route("/gerarPDF", methods = ["GET", "POST"])
def generatePDF():
    nomeEquipe = request.args.get("nomeEquipe")
    pdfPath = f"{UPLOAD_FOLDER_PDFS}{nomeEquipe}.pdf"
    pdf = canvas.Canvas(pdfPath, pagesize=A4)
    pdf.setTitle(f"{nomeEquipe}-notas")
    pdf = generateGrid(pdf=pdf,nomeEquipe=nomeEquipe)

    if pdf:
        pdf.save()
        return jsonify({'url': pdfPath}), 200
    return jsonify({'error': "Equipe Não encontrada"}), 404

def generateGrid(pdf, nomeEquipe):
    pathArquivo = f"{UPLOAD_EQUIPE}{nomeEquipe}.txt"
    if os.path.exists(pathArquivo):
        with open(pathArquivo, 'r', encoding='utf-8') as file:
            arquivo = file.readlines()
        
        json_notas = generateJson(jsonFile=arquivo)
        tupla_notas = generateTupla(jsonFile=json_notas)

        w, h = A4
        max_rows = 45

        margin_x = 30
        margin_y = 50
        padding = 15

        xList = [x + margin_x for x in [0, 100, 200, 300, 400, 530]]
        yList = [h - margin_y - i * padding for i in range(max_rows + 1)]

        for rows in grouper(tupla_notas, max_rows):
            rows = tuple(filter(bool, rows))
            pdf.grid(xList, yList[:len(rows) + 1])
            for y, row in zip(yList[:len(rows) + 1], rows):
                for x, cell in zip(xList, row):
                    pdf.drawString(x + 2, y - padding + 3, str(cell))
            pdf.showPage()
        
        return pdf

    return False

def generateJson(jsonFile):
    json_notas = {}
    for nota in jsonFile:
        nota = nota.strip().split(":", 1)

        chave, valor = nota[0], nota[1]
        chave = chave.strip().split("-")
        chave = chave[1]

        valor = valor.strip()

        if chave not in json_notas:
            json_notas[chave] = {"notas":[json.loads(valor)]}
        else:
            json_notas[chave]["notas"].append(json.loads(valor))

    json_notas = mediaNotas(jsonFile=json_notas)
    json_notas = ordenarJson(jsonFile=json_notas)

    return json_notas

def ordenarJson(jsonFile):
    jsonFile_ordenado = dict(sorted(jsonFile.items()))
    return jsonFile_ordenado

def mediaNotas(jsonFile):
    medias = {}

    for integrante, info in jsonFile.items():
        soma = {}
        contador = {}

        for notas in info["notas"]:
            for criterio, valor in notas.items():
                valor = int(valor)
                soma[criterio] = soma.get(criterio, 0) + valor
                contador[criterio] = contador.get(criterio, 0) + 1
        
        medias["medias"] = {
            criterio: round(soma[criterio]/contador[criterio], 1) for criterio in soma
        }

        jsonFile[integrante].update(medias)
    return jsonFile

def generateTupla(jsonFile):
    data = [("Nome", "Proatividade", "Autonomia", "Cooperação", "Entrega de Resultados")]

    for integrante, notaMedias in jsonFile.items():
        medias = []

        for criterio, valorMedia in notaMedias["medias"].items():
            medias.append(valorMedia)
        
        data.append((integrante, *medias))
    return data

def grouper(iteravel, n):
    args = [iter(iteravel)] * n
    return itertools.zip_longest(*args)

if __name__ == '__main__':
    app.run(debug=True, use_reloader=True, host="0.0.0.0")
